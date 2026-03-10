# --- GUI ---
import tkinter as tk
from tkinter import ttk, filedialog

# --- Plotting ---
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

# --- Numerical Analysis ---
import numpy as np
import pandas as pd   # ✅ Added so WatParser can use pd.DataFrame, pd.read_csv, etc.

# --- Excel Handling ---
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook
# --- Utilities ---
import os
import logging
from datetime import datetime, timedelta
import re
import sys

# WAT Data Automation Tool v1.1.1
# Author: Rose Anne Lafuente
# Licensed Electronics Engineer | Product Engineer II | Python Automation
#
# Description:
#   Automates .wat-to-Excel workflows for semiconductor Wafer Acceptance Test (WAT) data.
#   Generates structured sheets and capability plots for yield analysis and reporting.
#
# Features:
#   - Dynamic wafer sheet naming (Wafer 1~N, where N is fetched from the .wat file)
#   - Per Unit Data, Per Wafer, and Summary sheets with audit-ready formatting
#   - Cp, Cpk, Cpk Hi, and Cpk Lo statistics for process capability evaluation
#   - Interactive histogram viewer with ±3σ normal curve overlay
#   - Scrollable GUI logs with success/error messages for transparency
#   - Centralized error logging with 30-day cleanup
#   - Modular multi-class architecture (Parser, Builder, Visualizer, GUI Controller)
#   - GUI branding with sprout.ico and graceful fallback if icon not found
#
# Built with:
#   Python, Tkinter, OpenPyXL, Matplotlib, NumPy, Pandas

# --- Error Logger with 30-day cleanup ---
class ErrorLogger:
    """
    Centralized error logging for WAT Data Automation.

    - Creates timestamped error log files in a /logs directory.
    - Cleans up logs older than a defined retention period (default: 30 days).
    - Provides dual reporting: GUI-friendly messages and persistent log files.
    """

    def __init__(self, days_to_keep=30):
        self.log_dir = os.path.join(os.path.dirname(__file__), "logs")
        self.days_to_keep = days_to_keep
        self.log_file = None
        self.is_configured = False
        self.cleanup_old_logs()

    def setup_on_error(self):
        if not self.is_configured:
            os.makedirs(self.log_dir, exist_ok=True)
            self.log_file = os.path.join(
                self.log_dir,
                f"wat_data_automation_error_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
            )
            logging.basicConfig(
                filename=self.log_file,
                level=logging.ERROR,
                format="%(asctime)s - %(levelname)s - %(message)s"
            )
            self.is_configured = True
        return self.log_file

    def log_error(self, msg: str):
        """Log an error message to file and return a GUI-friendly string."""
        self.setup_on_error()
        logging.error(msg)
        return f"❌ {msg}"

    def log_success(self, msg: str):
        """Return a GUI-friendly success message (no file logging)."""
        return f"✅ {msg}"

    def cleanup_old_logs(self):
        """Delete log files older than the retention period."""
        if not os.path.exists(self.log_dir):
            return
        cutoff = datetime.now() - timedelta(days=self.days_to_keep)
        for fname in os.listdir(self.log_dir):
            fpath = os.path.join(self.log_dir, fname)
            if os.path.isfile(fpath) and fname.startswith("wat_data_automation_error_log_"):
                try:
                    timestamp_str = fname.replace("wat_data_automation_error_log_", "").replace(".txt", "")
                    dt = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                    if dt < cutoff:
                        os.remove(fpath)
                except Exception:
                    continue
                
def resource_path(relative_path):
    """Get absolute path to resource, works for dev and PyInstaller"""
    try:
        base_path = sys._MEIPASS  # PyInstaller temp folder
    except Exception:
        base_path = os.path.abspath(".")
    return os.path.join(base_path, relative_path)

# --- Shared Helper Functions (GLOBAL) ---
def write_text_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=str(value) if value not in (None, '') else "")
    cell.number_format = '@'
    return cell

def write_number_cell(ws, row, col, value, round_digits=3):
    try:
        if value not in (None, ''):
            num = float(value)
            if num.is_integer():
                ws.cell(row=row, column=col, value=int(num))
            else:
                rounded_val = round(num, round_digits)
                cell = ws.cell(row=row, column=col, value=rounded_val)
                cell.number_format = '0.000'
        else:
            ws.cell(row=row, column=col, value=value)
    except Exception:
        ws.cell(row=row, column=col, value=value)

def autofit_columns(ws, min_col=1, min_row=1):
    for col in ws.iter_cols(min_col=min_col, min_row=min_row, max_row=ws.max_row):
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 3

def apply_borders(ws, min_row=1, min_col=1):
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))
    last_row = 0
    for row in range(1, ws.max_row + 1):
        if any(ws.cell(row=row, column=col).value not in (None, "")
               for col in range(1, ws.max_column + 1)):
            last_row = row
    for row in ws.iter_rows(min_row=min_row, max_row=last_row,
                            min_col=min_col, max_col=ws.max_column):
        for cell in row:
            cell.border = thin_border

def find_param_rows(ws):
    waf_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                for cell in row if str(cell.value).strip().upper() == "WAF"]
    id_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
               for cell in row if str(cell.value).strip().upper() == "ID"]
    spechi_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                   for cell in row if str(cell.value).strip().upper() == "SPEC HI"]
    speclo_rows = [cell.row for row in ws.iter_rows(min_col=1, max_col=1)
                   for cell in row if str(cell.value).strip().upper() == "SPEC LO"]
    return waf_rows, id_rows, spechi_rows, speclo_rows

def build_param_mapping(ws, waf_rows, id_rows, spechi_rows, speclo_rows):
    param_mapping = {}
    min_len = min(len(waf_rows), len(id_rows), len(spechi_rows), len(speclo_rows))
    for i in range(min_len):
        waf_row, id_row, hi_row, lo_row = waf_rows[i], id_rows[i], spechi_rows[i], speclo_rows[i]
        param_headers = [ws.cell(row=waf_row, column=col).value
                         for col in range(3, ws.max_column + 1)]
        for idx, param in enumerate(param_headers):
            if param is None:
                continue
            col = 3 + idx
            unit = ws.cell(row=id_row, column=col).value
            spec_hi = ws.cell(row=hi_row, column=col).value
            spec_lo = ws.cell(row=lo_row, column=col).value
            param_mapping[str(param).strip()] = (spec_hi, spec_lo, unit)
            if isinstance(spec_hi, (int, float)) and isinstance(spec_lo, (int, float)):
                param_mapping[f"{param}_POS"] = (abs(spec_lo), abs(spec_hi), unit)
    return param_mapping

def extract_site_values( ws, start_row=6, col="B"):
    """Extract unique SITE values from a worksheet starting at a given row."""
    site_values = []
    row = start_row
    while True:
        cell_val = ws[f"{col}{row}"].value
        if cell_val is None or str(cell_val).strip().upper() in ("AVERAGE", "STD", "SPEC"):
            break
        clean_site = str(cell_val).replace('-', '').strip()
        if clean_site and clean_site not in site_values:
            site_values.append(clean_site)
        row += 1
    return site_values
    
# --- WatParser Class ---
class WatParser:
    """
    Handles parsing of raw .wat files into structured Excel workbooks.

    Responsibilities:
    - Extracts metadata (TYPE NO, PROCESS, PCM SPEC, LOT ID, DATE, QTY).
    - Converts tabular WAT data into a Pandas DataFrame.
    - Writes metadata and data into an initial Excel workbook.
    """
    def __init__(self, logger):
        self.logger = logger

    
    def parse_file(self, wat_file, excel_filename):
        try:
            # Step 1: Read full content
            with open(wat_file, 'r') as file:
                content_text = file.read()
                file.seek(0)
                content_lines = file.readlines()

            # Step 2: Extract metadata using regex
            wat_attached = "W.A.T DATA ATTACHED"
            type_no = re.search(r'TYPE NO\s*:(\S+)', content_text).group(1)
            process = re.search(r'PROCESS\s*:(\S+)', content_text).group(1)
            pcm_spec = re.search(r'PCM SPEC\s*:(\S+)', content_text).group(1)
            qty = re.search(r'QTY\s*:(.+)', content_text).group(1).strip()
            lot_id = re.search(r'LOT ID\s*:(\S+)', content_text).group(1)
            date = re.search(r'DATE\s*:(\S+)', content_text).group(1)

            # Step 3: Extract table content
            header = content_lines[3].strip().split()
            data = [line.strip().split() for line in content_lines[4:] if line.strip()]
            df = pd.DataFrame(data, columns=header)

            # Step 4: Write to Excel
            wb = Workbook()
            ws = wb.active
            clean_qty = qty.replace("pcs", "").strip()
            ws.title = f"Wafer 1~{clean_qty}"
            ws.sheet_view.showGridLines = False

            # Metadata cells
            data_to_excel = {
                (1, 5): wat_attached,
                (2, 1): "TYPE NO :", (2, 2): type_no,
                (2, 5): "PROCESS :", (2, 6): process,
                (2, 8): "PCM SPEC :", (2, 9): pcm_spec,
                (2, 11): "QTY :", (2, 12): qty,
                (3, 1): "LOT ID :", (3, 2): lot_id,
                (3, 5): "DATE :", (3, 6): date,
            }
            for (row, col), value in data_to_excel.items():
                ws.cell(row=row, column=col, value=value)

            # Write DataFrame starting at row 4
            start_row = 4
            for r_idx, row_values in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
                if str(row_values[0]).strip().upper() == "AVERAGE":
                    write_text_cell(ws, r_idx, 1, "AVERAGE")
                    write_text_cell(ws, r_idx, 2, "")
                    for c_idx, value in enumerate(row_values[1:], start=3):
                        write_number_cell(ws, r_idx, c_idx, value)
                elif (
                    len(row_values) >= 3 and
                    isinstance(row_values[0], str) and
                    row_values[0].upper().startswith(("STD", "SPEC")) and
                    isinstance(row_values[1], str)
                ):
                    combined_label = f"{row_values[0]} {row_values[1]}"
                    write_text_cell(ws, r_idx, 1, combined_label)
                    for c_idx, value in enumerate(row_values[2:], start=3):
                        write_number_cell(ws, r_idx, c_idx, value)
                else:
                    for c_idx, value in enumerate(row_values, start=1):
                        if c_idx <= 2:  # WAF and SITE columns → text
                            write_number_cell(ws, r_idx, c_idx, value)
                        else:           # Parameters → numbers
                            write_number_cell(ws, r_idx, c_idx, value)

            # --- Step 5: Auto-fit column widths (from Column C, Row 4) ---
            autofit_columns(ws, min_col=3, min_row=4)

            # --- Step 6: Save file ---

            wb.save(excel_filename)
            return type_no, process, pcm_spec, lot_id, date, clean_qty

        except Exception as e:
            raise RuntimeError(self.logger.log_error(f"Error in WatParser: {e}"))

# --- WorkbookBuilder Class ---
class WorkbookBuilder:
    """
    Builds and formats Excel sheets for WAT Data Automation.

    Responsibilities:
    - Adds 'Per Unit Data' sheet with wafer/site breakdown.
    - Adds 'Per Wafer' sheet with site-level statistics and spec mapping.
    - Generates statistical summary sheets with Cp/Cpk metrics.
    - Provides helper functions for formatting, borders, and autofit.
    """
    def __init__(self, logger):
        self.logger = logger

    # --- Per Unit Data ---
    def add_per_unit_data(self, excel_filename, clean_qty):
        try:
            wb = load_workbook(excel_filename)
            main_ws = wb[wb.sheetnames[0]]

            # Metadata
            type_no = main_ws["B2"].value
            process = main_ws["F2"].value
            pcm_spec = main_ws["I2"].value
            lot_id = main_ws["B3"].value
            date = main_ws["F3"].value

            # Extract SITE values using global helper
            site_values = extract_site_values(main_ws, start_row=6, col="B")

            # Wafer IDs
            wafer_ids = []
            for wafer_num in range(1, int(clean_qty) + 1):
                wafer_label = f"TT_#{str(wafer_num).zfill(2)}"
                for site in site_values:
                    wafer_ids.append((wafer_label, site))

            # Create sheet
            unit_ws = wb.create_sheet(title="per Unit Data")
            headers = ["TYPE", "PROCESS", "SPEC", "LOT", "DATE", "Wafer", "Site"]
            unit_ws.append(headers)
            for wafer, site in wafer_ids:
                site_number = site_values.index(site) + 1
                unit_ws.append([type_no, process, pcm_spec, lot_id, date, wafer, site_number])

            # Parameter blocks
            waf_rows, id_rows, spechi_rows, speclo_rows = find_param_rows(main_ws)

            current_header_col = 8
            for waf_row in waf_rows:
                # Iterate over row cells correctly
                param_headers = [cell.value for cell in main_ws[waf_row] if cell.column >= 3]

                # Write headers
                for i, header in enumerate(param_headers):
                    unit_ws.cell(row=1, column=current_header_col + i, value=header)

                start_value_row = waf_row + 2
                current_data_row = 2
                while True:
                    label = main_ws[f"A{start_value_row}"].value
                    if label in (None, "", "AVERAGE", "STD", "SPEC"):
                        break
                    for i, header in enumerate(param_headers):
                        value = main_ws.cell(row=start_value_row, column=3 + i).value
                        write_number_cell(unit_ws, current_data_row, current_header_col + i, value)
                    start_value_row += 1
                    current_data_row += 1
                current_header_col += len(param_headers)

            # Add _POS columns
            param_headers = []
            col = 8
            while True:
                header = unit_ws.cell(row=1, column=col).value
                if header is None or header == "":
                    break
                param_headers.append((col, header))
                col += 1
            next_col = col
            created_pos_columns = {}
            for col_idx, header in param_headers:
                for row in range(2, unit_ws.max_row + 1):
                    val = unit_ws.cell(row=row, column=col_idx).value
                    if isinstance(val, (int, float)) and val < 0:
                        pos_header = f"{header}_POS"
                        if pos_header not in created_pos_columns:
                            unit_ws.cell(row=1, column=next_col, value=pos_header)
                            created_pos_columns[pos_header] = next_col
                            next_col += 1
                        unit_ws.cell(row=row, column=created_pos_columns[pos_header], value=abs(val))

            # Formatting
            unit_ws.sheet_view.showGridLines = False
            header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            header_font = Font(bold=True)

            # Style only non-empty headers and track the last one
            last_header_col = 0
            for cell in unit_ws[1]:
                if cell.value not in (None, ""):   # ✅ skip empty headers
                    cell.font = header_font
                    cell.fill = header_fill
                    last_header_col = cell.column  # track last non-empty header col

            # Apply helpers once
            apply_borders(unit_ws)
            autofit_columns(unit_ws)

            # ✅ Cleanup: remove borders and reset widths for trailing empty columns
            for row in unit_ws.iter_rows(min_row=1, max_row=unit_ws.max_row, min_col=last_header_col+1, max_col=unit_ws.max_column):
                for cell in row:
                    cell.border = Border()  # clear border

            for col in range(last_header_col+1, unit_ws.max_column+1):
                col_letter = get_column_letter(col)
                unit_ws.column_dimensions[col_letter].width = 0  # collapse autofit width

            wb.save(excel_filename)
            return self.logger.log_success("New 'Per Unit Data' sheet with formatting created.")
        except Exception as e:
            raise RuntimeError(self.logger.log_error(f"Error in add_per_unit_data: {e}"))

    # --- Add Per Wafer Data ---
    def add_per_wafer_data(self, excel_filename, clean_qty, site_values, type_no, process, pcm_spec, lot_id, date):
        try:
            wb = load_workbook(excel_filename)
            main_ws = wb["per Unit Data"]

            # Collect parameters
            Param = [main_ws.cell(row=1, column=col).value
                     for col in range(8, main_ws.max_column + 1)
                     if main_ws.cell(row=1, column=col).value]

            # Extract SITE values using global helper
            data_ws = wb[wb.sheetnames[0]]
            site_values = extract_site_values(data_ws, start_row=6, col="B")

            site_start_col = 8  # Column H
            site_end_col = site_start_col + len(site_values) - 1

            per_Wafer_ws = wb.create_sheet(title="per Wafer")

            # Header row 1: General headers + merged 'Site' + stat headers
            headers_row1 = ["TYPE", "PROCESS", "SPEC", "LOT", "DATE", "WAF_ID", "Parameter"]
            headers_row1 += ["Site"] + [""] * (len(site_values) - 1)
            headers_row1 += ["AVERAGE", "STD_DEV", "SPEC HI", "SPEC LO", "Unit"]
            per_Wafer_ws.append(headers_row1)

            # Write site values in row 2
            for idx, site in enumerate(site_values, start=8):
                write_number_cell(per_Wafer_ws, row=2, col=idx, value=site)

            # Merge metadata headers vertically from A1 to G1
            for col_num in range(1, 7 + 1):
                col_letter = get_column_letter(col_num)
                per_Wafer_ws.merge_cells(f"{col_letter}1:{col_letter}2")
                per_Wafer_ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

            # Merge 'Site' group
            if site_values:
                site_merge_range = f"{get_column_letter(site_start_col)}1:{get_column_letter(site_end_col)}1"
                per_Wafer_ws.merge_cells(site_merge_range)
                per_Wafer_ws.cell(row=1, column=site_start_col).value = "Site"
                per_Wafer_ws.cell(row=1, column=site_start_col).alignment = Alignment(horizontal="center", vertical="center")

            # Merge stat headers vertically from row 1 to row 2
            stat_headers = ["AVERAGE", "STD_DEV", "SPEC HI", "SPEC LO", "Unit"]
            stat_start_col = site_end_col + 1
            for i, header in enumerate(stat_headers):
                col_index = stat_start_col + i
                col_letter = get_column_letter(col_index)
                per_Wafer_ws.merge_cells(f"{col_letter}1:{col_letter}2")
                per_Wafer_ws[f"{col_letter}1"].value = header
                per_Wafer_ws[f"{col_letter}1"].alignment = Alignment(horizontal="center", vertical="center")

            # Add wafer rows
            for wafer_num in range(1, int(clean_qty) + 1):
                wafer_id = f"TT_#{str(wafer_num).zfill(2)}"
                for param in Param:
                    per_Wafer_ws.append([type_no, process, pcm_spec, lot_id, date, wafer_id, param])

            # Copy and transpose values from "per Unit Data"
            unit_ws = wb["per Unit Data"]
            wafer_col = 6  # Column F (Wafer)
            param_start_col = 8  # Column H
            param_end_col = unit_ws.max_column

            # Get all unique wafer IDs
            wafer_ids = []
            for row in range(2, unit_ws.max_row + 1):
                wafer_val = unit_ws.cell(row=row, column=wafer_col).value
                if wafer_val and wafer_val not in wafer_ids:
                    wafer_ids.append(wafer_val)

            for wafer in wafer_ids:
                wafer_rows = [row for row in range(2, unit_ws.max_row + 1)
                              if unit_ws.cell(row=row, column=wafer_col).value == wafer]
                if not wafer_rows:
                    continue
                values_matrix = [[unit_ws.cell(row=row, column=col).value
                                  for col in range(param_start_col, param_end_col + 1)]
                                 for row in wafer_rows]
                if values_matrix:
                    transposed = list(map(list, zip(*values_matrix)))
                    for row_idx in range(3, per_Wafer_ws.max_row + 1):
                        if per_Wafer_ws.cell(row=row_idx, column=6).value == wafer:
                            for t_row_idx, t_row in enumerate(transposed):
                                for t_col_idx, val in enumerate(t_row):
                                    cell = per_Wafer_ws.cell(row=row_idx + t_row_idx,
                                                             column=8 + t_col_idx,
                                                             value=round(val, 3) if isinstance(val, (int, float)) else val)
                                    if isinstance(val, (int, float)):
                                        cell.number_format = '0.000'
                            break

            # Find last row with a parameter
            last_param_row = 0
            for row in range(3, per_Wafer_ws.max_row + 1):
                if per_Wafer_ws.cell(row=row, column=7).value not in (None, ""):
                    last_param_row = row

            # Add formulas only up to that row
            for row in range(3, last_param_row + 1):
                avg_col = site_end_col + 1
                stdev_col = site_end_col + 2

                avg_cell = per_Wafer_ws.cell(row=row, column=avg_col)
                stdev_cell = per_Wafer_ws.cell(row=row, column=stdev_col)

                avg_cell.value = f"=AVERAGE({get_column_letter(site_start_col)}{row}:{get_column_letter(site_end_col)}{row})"
                stdev_cell.value = f"=STDEV({get_column_letter(site_start_col)}{row}:{get_column_letter(site_end_col)}{row})"

                avg_cell.number_format = '0.000'
                stdev_cell.number_format = '0.000'

            # --- Inject Spec Mapping ---
            waf_rows, id_rows, spechi_rows, speclo_rows = find_param_rows(data_ws)
            param_mapping = build_param_mapping(data_ws, waf_rows, id_rows, spechi_rows, speclo_rows)

            # Fill Spec HI, Spec LO, Unit columns
            for row in range(2, per_Wafer_ws.max_row + 1):
                param_name = per_Wafer_ws.cell(row=row, column=7).value
                if param_name in param_mapping:
                    spec_hi, spec_lo, unit = param_mapping[param_name]
                    per_Wafer_ws.cell(row=row, column=site_end_col + 3, value=spec_hi)
                    per_Wafer_ws.cell(row=row, column=site_end_col + 4, value=spec_lo)
                    per_Wafer_ws.cell(row=row, column=site_end_col + 5, value=unit)

            # --- Format Header Rows of per Wafer sheet ---
            header_fill = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
            header_font = Font(bold=True)

            for row_idx in [1, 2]:
                for cell in per_Wafer_ws[row_idx]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            # Auto-fit and borders using global helpers
            per_Wafer_ws.sheet_view.showGridLines = False
            autofit_columns(per_Wafer_ws)
            apply_borders(per_Wafer_ws)
            # ✅ Recalculate widths for stat/formula columns dynamically
            stat_headers = ["AVERAGE", "STD_DEV", "SPEC HI", "SPEC LO", "Unit"]

            for header in stat_headers:
                for cell in per_Wafer_ws[1]:
                    if cell.value == header:
                        col_idx = cell.column
                        col_letter = get_column_letter(col_idx)

                        # Start with header length
                        max_len = len(str(header))

                        # Measure actual values (numeric formatted to 3 decimals)
                        for row in range(3, per_Wafer_ws.max_row + 1):
                            val = per_Wafer_ws.cell(row=row, column=col_idx).value
                            if val not in (None, ""):
                                if isinstance(val, (int, float)):
                                    text = f"{val:.3f}"
                                else:
                                    text = str(val)
                                max_len = max(max_len, len(text))

                        # Apply dynamic width with padding
                        per_Wafer_ws.column_dimensions[col_letter].width = max_len + 2


            wb.save(excel_filename)
            return self.logger.log_success("Per Wafer sheet with Spec Mapping added.")
        except Exception as e:
            raise RuntimeError(self.logger.log_error(f"Error in add_per_wafer_data: {e}"))
    
    # --- Generate Statistical Summary ---
    def run_summary(self, excel_filename):
        if not excel_filename:
            return self.logger.log_error("⚠️ No Excel file selected yet.")
        try:
            wb = load_workbook(excel_filename)
        except Exception as e:
            raise RuntimeError(self.logger.log_error(f"❌ Could not open {excel_filename}. Error: {e}"))

        unit_ws = wb["per Unit Data"]
        per_Wafer_ws = wb["per Wafer"]
        data_ws = wb[wb.sheetnames[0]]
        lot_id = data_ws["B3"].value or "UNKNOWN"

        # --- Detect site_end_col dynamically from row 2 ---
        site_start_col = 8  # Column H
        site_values = []
        for col in range(site_start_col, per_Wafer_ws.max_column + 1):
            val = per_Wafer_ws.cell(row=2, column=col).value
            if val not in (None, ""):
                site_values.append(val)
        site_end_col = site_start_col + len(site_values) - 1

        # --- Define styles ---
        fill_green = PatternFill(start_color="FF00823B", end_color="FF00823B", fill_type="solid")
        fill_blue  = PatternFill(start_color="FF156082", end_color="FF156082", fill_type="solid")
        white_bold = Font(bold=True, color="FFFFFF")
        thin_border = Border(left=Side(style="thin"), right=Side(style="thin"),
                             top=Side(style="thin"), bottom=Side(style="thin"))

        # --- Create summary sheet ---
        param_map_ws = wb.create_sheet(title=f"{lot_id}_TTTT_WAT_Summary")
        param_map_ws.sheet_view.showGridLines = False

        # Merge headers
        param_map_ws.merge_cells("B2:B3")
        param_map_ws.merge_cells("C2:C3")
        param_map_ws.merge_cells("D2:D3")
        param_map_ws.merge_cells("E2:I2")

        # Title cell
        param_map_ws["E2"] = f"{lot_id}_TTTT_Summary"
        param_map_ws["E2"].alignment = Alignment(horizontal="center", vertical="center")
        param_map_ws["E2"].fill = fill_green
        param_map_ws["E2"].font = white_bold

        # Header labels
        param_map_ws["B2"] = "Parameter"
        param_map_ws["C2"] = "SPEC HI"
        param_map_ws["D2"] = "SPEC LO"
        param_map_ws["E3"] = "MEAN"
        param_map_ws["F3"] = "STDEV"
        param_map_ws["G3"] = "CPK"
        param_map_ws["H3"] = "CPK Hi"
        param_map_ws["I3"] = "CPK Lo"

        # Style B–D headers
        for row in [2, 3]:
            for col in range(2, 5):
                cell = param_map_ws.cell(row=row, column=col)
                cell.fill = fill_blue
                cell.font = white_bold
                cell.alignment = Alignment(horizontal="center", vertical="center")

        # Style E–I headers
        for col in range(5, 10):
            cell = param_map_ws.cell(row=3, column=col)
            cell.fill = fill_green
            cell.font = white_bold
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # --- Build mapping of Parameter : SPEC HI : SPEC LO ---
        param_map = {}
        for row in range(3, per_Wafer_ws.max_row + 1):
            param = per_Wafer_ws.cell(row=row, column=7).value
            spec_hi = per_Wafer_ws.cell(row=row, column=site_end_col + 3).value
            spec_lo = per_Wafer_ws.cell(row=row, column=site_end_col + 4).value

            if param:
                param = str(param).strip()
                if param not in param_map:
                    param_map[param] = {"SPEC HI": spec_hi, "SPEC LO": spec_lo}

        # --- Compute MEAN and STDEV from per Unit Data ---
        last_row = unit_ws.max_row
        for col in range(8, unit_ws.max_column + 1):  # start at H
            param_name = unit_ws.cell(row=1, column=col).value
            if not param_name:
                continue

            values = []
            for row in range(2, last_row + 1):
                val = unit_ws.cell(row=row, column=col).value
                if val is not None:
                    try:
                        values.append(float(val))
                    except ValueError:
                        pass

            if values:
                arr = np.array(values, dtype=float)
                mean_val = round(float(np.mean(arr)), 3)
                stdev_val = round(float(np.std(arr, ddof=1)), 3)  # STDEV.S
                if param_name.strip() in param_map:
                    param_map[param_name.strip()]["MEAN"] = mean_val
                    param_map[param_name.strip()]["STDEV"] = stdev_val

                    # --- Compute CPK metrics ---
                    spec_hi = param_map[param_name.strip()]["SPEC HI"]
                    spec_lo = param_map[param_name.strip()]["SPEC LO"]

                    if spec_hi is not None and spec_lo is not None and stdev_val != 0:
                        cpk1 = round((spec_hi - mean_val) / (3 * stdev_val), 3)
                        cpk2 = round((mean_val - spec_lo) / (3 * stdev_val), 3)
                        param_map[param_name.strip()]["CPK"] = min(cpk1, cpk2)
                        param_map[param_name.strip()]["CPK Hi"] = max(cpk1, cpk2)
                        param_map[param_name.strip()]["CPK Lo"] = min(cpk1, cpk2)
                    else:
                        param_map[param_name.strip()]["CPK"] = "na"
                        param_map[param_name.strip()]["CPK Hi"] = "na"
                        param_map[param_name.strip()]["CPK Lo"] = "na"

        # --- Paste into summary sheet ---
        row_num = 4
        for param, specs in param_map.items():
            for col_idx, key in enumerate(
                ["Parameter", "SPEC HI", "SPEC LO", "MEAN", "STDEV", "CPK", "CPK Hi", "CPK Lo"], start=2
            ):
                val = specs.get(key) if key != "Parameter" else param
                cell = param_map_ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = thin_border
            row_num += 1

        # Apply borders to header area too
        for row in param_map_ws.iter_rows(min_row=2, max_row=row_num-1, min_col=2, max_col=9):
            for cell in row:
                cell.border = thin_border

        # ✅ Use global helper instead of self.apply_borders
        apply_borders(param_map_ws, min_row=2, min_col=2)

        wb.save(excel_filename)
        return self.logger.log_success("Summary sheet successfully created.")
        
# --- HistogramGUI Class ---
class HistogramGUI:
    def __init__(self, root, excel_filename, logger):
        self.root = root
        self.excel_filename = excel_filename
        self.logger = logger

        self.root.title("Histogram Viewer")
        self.root.geometry("960x600")  # balanced window size
        try:
            self.root.iconbitmap("sprout.ico")  # Add branding icon
        except Exception:
            pass  # Ignore if icon not found
        
        # --- Left frame for plot ---
        left_frame = tk.Frame(self.root)
        left_frame.pack(side="left", fill="both", expand=True)

        self.fig, self.ax = plt.subplots(figsize=(8, 5))
        # Adjust subplot margins to leave space on the right
        self.fig.subplots_adjust(right=0.75)

        self.canvas = FigureCanvasTkAgg(self.fig, master=left_frame)
        self.canvas.get_tk_widget().pack(fill="both", expand=True)

        # Add the toolbar (includes Save, Zoom, Pan, etc.)
        toolbar = NavigationToolbar2Tk(self.canvas, left_frame)
        # Replace the default save button behavior
        toolbar.save_figure = self.save_plot
        toolbar.update()
        self.canvas._tkcanvas.pack(fill="both", expand=True)

        # Stats text embedded in the figure (white box, right side)
        self.stats_text_obj = self.fig.text(
            0.78, 0.87, "", fontsize=9,
            va="top", ha="left", family="monospace",
            bbox=dict(facecolor="white", alpha=0.8, edgecolor="black")
        )

        # --- Right frame for parameter listbox ---
        right_frame = tk.Frame(self.root, bg="#f5f5f5", width=250)
        right_frame.pack(side="right", fill="y", padx=10, pady=10)
        right_frame.pack_propagate(False)  # keep fixed width

        # Label stretches across the frame
        tk.Label(
            right_frame,
            text="Select Parameter",
            font=("Segoe UI", 10, "bold"),
            anchor="w"
        ).pack(fill="x", pady=5)

        # Frame for listbox + scrollbar with fixed height
        listbox_frame = tk.Frame(right_frame, height=400)  # adjust height here
        listbox_frame.pack(fill="x", pady=5)
        listbox_frame.pack_propagate(False)  # prevent shrinking to children

        # Listbox inside fixed-height frame
        self.param_listbox = tk.Listbox(listbox_frame)
        self.param_listbox.pack(side="left", fill="both", expand=True)

        scrollbar = tk.Scrollbar(listbox_frame, orient="vertical")
        scrollbar.pack(side="right", fill="y")

        self.param_listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.param_listbox.yview)


        # Bind click event to listbox
        self.param_listbox.bind("<<ListboxSelect>>", self.on_select)

        # Populate listbox with parameters
        self.parameters = self.load_parameters()
        for param in self.parameters:
            self.param_listbox.insert("end", param)

        # Initial plot
        if self.parameters:
            self.plot_hist(self.parameters[0])

    def load_parameters(self):
        """Extract parameter names from Column G of 'per Wafer' sheet."""
        try:
            wb = load_workbook(self.excel_filename, data_only=True)
            ws = wb["per Wafer"]

            param_col = 7  # Column G
            params = []
            for row in range(2, ws.max_row + 1):
                val = ws.cell(row=row, column=param_col).value
                if val:
                    params.append(val)
            return sorted(set(params))
        except Exception as e:
            self.logger.log_error(f"Failed to load parameters: {e}")
            return []

    def get_values(self, param):
        """Fetch values, specs, and unit for a parameter from per Wafer sheet."""
        try:
            wb = load_workbook(self.excel_filename, data_only=True)
            ws = wb["per Wafer"]

            param_col = 7
            site_start_col = 8
            site_end_col = ws.max_column - 5

            values, spec_hi, spec_lo, unit = [], None, None, None
            for row in range(2, ws.max_row + 1):
                if ws.cell(row=row, column=param_col).value == param:
                    row_values = [
                        ws.cell(row=row, column=col).value
                        for col in range(site_start_col, site_end_col + 1)
                        if isinstance(ws.cell(row=row, column=col).value, (int, float))
                    ]
                    values.extend(row_values)
                    if spec_hi is None:
                        spec_hi = ws.cell(row=row, column=site_end_col + 3).value
                        spec_lo = ws.cell(row=row, column=site_end_col + 4).value
                        unit    = ws.cell(row=row, column=site_end_col + 5).value
            return values, spec_hi, spec_lo, unit
        except Exception as e:
            self.logger.log_error(f"Error fetching values: {e}")
            return [], None, None, None
        
    def plot_hist(self, param):
        try:
            values, spec_hi, spec_lo, unit = self.get_values(param)
            if not values:
                return

            mean_val = np.mean(values)
            stdev_val = np.std(values, ddof=1)
            min_val, max_val = np.min(values), np.max(values)
            count_val = len(values)

            # Capability stats
            if spec_hi is not None and spec_lo is not None and stdev_val != 0:
                cpk_hi = (spec_hi - mean_val) / (3 * stdev_val)
                cpk_lo = (mean_val - spec_lo) / (3 * stdev_val)
                cpk = min(cpk_hi, cpk_lo)
                cp = (spec_hi - spec_lo) / (6 * stdev_val)
            else:
                cpk_hi = cpk_lo = cpk = cp = None

            cp_str = f"{cp:.3f}" if cp is not None else "na"
            cpk_str = f"{cpk:.3f}" if cpk is not None else "na"
            cpk_hi_str = f"{cpk_hi:.3f}" if cpk_hi is not None else "na"
            cpk_lo_str = f"{cpk_lo:.3f}" if cpk_lo is not None else "na"

            # Clear and redraw histogram
            self.ax.clear()
            self.ax.hist(values, bins=20, color="#4F81BD", edgecolor="black", alpha=0.7, density=True)

            # Normal curve
            x = np.linspace(mean_val - 4*stdev_val, mean_val + 4*stdev_val, 400)
            y = (1 / (stdev_val * np.sqrt(2 * np.pi))) * np.exp(-0.5 * ((x - mean_val) / stdev_val) ** 2)
            self.ax.plot(x, y, color="orange", linewidth=2, label="Normal Curve")

            # Vertical lines
            self.ax.axvline(mean_val, color="green", linestyle="--", linewidth=2, label=f"Mean = {mean_val:.3f}")
            if spec_lo is not None:
                self.ax.axvline(spec_lo, color="red", linestyle="--", linewidth=2, label=f"LSL = {spec_lo}")
            if spec_hi is not None:
                self.ax.axvline(spec_hi, color="red", linestyle="--", linewidth=2, label=f"USL = {spec_hi}")
            self.ax.axvline(mean_val - 3*stdev_val, color="gray", linestyle=":", label="-3σ")
            self.ax.axvline(mean_val + 3*stdev_val, color="gray", linestyle=":", label="+3σ")

            self.ax.set_title(f"{param}")
            self.ax.set_xlabel(f"{unit}" if unit else "Value")
            self.ax.set_ylabel("Count")
            self.ax.legend()
            self.ax.grid(axis="y", linestyle="--", alpha=0.7)

            # Update stats text
            stats_text = f"""
Basic Statistics
Count: {count_val}
Min: {min_val:.3f}
Max: {max_val:.3f}
Mean: {mean_val:.3f}
Std Dev: {stdev_val:.3f}

Specifications
LSL: {spec_lo}
USL: {spec_hi}
Target: {mean_val:.3f}

Capability Statistics
Cp: {cp_str}
Cpk: {cpk_str}
Cpk Hi: {cpk_hi_str}
Cpk Lo: {cpk_lo_str}
"""
            self.stats_text_obj.set_text(stats_text)

            self.canvas.draw_idle()
            self.logger.log_success(f"Histogram plotted for {param}")

        except Exception as e:
            self.logger.log_error(f"Error generating histogram: {e}")

    def on_select(self, event):
        selection = self.param_listbox.curselection()
        if selection:
            param = self.param_listbox.get(selection[0])
            self.plot_hist(param)
            
    def save_plot(self):
        # Get the selected parameter from the listbox
        selection = self.param_listbox.curselection()
        if not selection:
            self.logger.log_error("No parameter selected to save.")
            return

        param = self.param_listbox.get(selection[0])

        # Save the figure with the parameter name
        filename = f"{param}_histogram.png"
        self.fig.savefig(filename, dpi=300, bbox_inches="tight")

        self.logger.log_success(f"Plot saved as {filename}")
# --- GuiController Class ---
class GuiController:
    """
    Tkinter-based GUI controller for WAT Data Automation.

    Responsibilities:
    - Provides file selection, action buttons, and status display.
    - Delegates parsing, workbook building, and visualization to modular classes.
    - Handles GUI branding with sprout.ico (graceful fallback if missing).
    - Routes all errors through ErrorLogger for consistency.
    """

    def __init__(self, root):
        self.root = root
        self.logger = ErrorLogger()
        self.parser = WatParser(self.logger)
        self.builder = WorkbookBuilder(self.logger)
        self.path_var = tk.StringVar()

        # Window setup
        self.root.title("WAT Data Automation v1.1.1")
        self.root.geometry("600x500")
        try:
            self.root.iconbitmap("sprout.ico")  # Add branding icon
        except Exception:
            pass  # Ignore if icon not found

        # Theme
        self.bg_color = "#f5f5f5"
        self.fg_color = "#222222"
        self.btn_bg = "#e0e0e0"
        self.btn_active = "#BEE395"
        self.root.configure(bg=self.bg_color)

        # Title
        title_frame = tk.Frame(self.root, bg=self.bg_color)
        title_frame.pack(pady=(10, 0))
        tk.Label(title_frame, text="WAT Data Automation",
                 font=("Meiryo", 12, "bold"), fg="darkblue", bg=self.bg_color).pack(side="left")
        tk.Label(title_frame, text=" v1.1.1",
                 font=("Meiryo", 12, "italic"), fg="darkblue", bg=self.bg_color).pack(side="left")
        tk.Label(self.root, text="Developed by Rose Anne Lafuente | 2026",
                 font=("Arial", 7, "italic"), fg="gray", bg=self.bg_color).pack(pady=(0, 10))

        self.create_file_selection_frame()
        self.create_action_buttons()
        self.create_status_box()
        self.create_exit_button()
        # Bind the window close event
        self.root.protocol("WM_DELETE_WINDOW", self.on_close)

    def on_close(self):
        # Clean up matplotlib canvas
        try:
            self.canvas.get_tk_widget().destroy()
            plt.close(self.fig)
        except Exception:
            pass

        # Destroy the root window
        self.root.quit()
        self.root.destroy()

        # Force process exit
        sys.exit(0)

    # --- File selection ---
    def create_file_selection_frame(self):
        f = tk.LabelFrame(self.root, text="WAT File Selection", padx=10, pady=10,
                          bd=2, relief="groove", font=("Segoe UI", 10, "bold"))
        f.pack(fill="x", padx=15, pady=10)

        tk.Label(f, text="Select WAT File:").pack(side="left", padx=(0, 10), pady=5)
        tk.Entry(f, textvariable=self.path_var, width=40, bg="white", fg="black").pack(side="left", fill="x", expand=True)
        tk.Button(f, text="Browse", width=15,
                  command=self.browse_file,
                  bg=self.btn_bg, fg=self.fg_color, activebackground=self.btn_active).pack(side="right", padx=10)

    def browse_file(self):
        file_path = filedialog.askopenfilename(title="Select WAT File", filetypes=[("WAT files", "*.wat")])
        if file_path:
            self.path_var.set(file_path)
            self.show_status(self.logger.log_success(f"📂 Selected file: {file_path}"))

    def create_action_buttons(self):
        action_frame = tk.Frame(self.root, bg=self.bg_color, height=70)
        action_frame.pack(pady=10, anchor="center")
        action_frame.pack_propagate(False)
        for col in range(3):
            action_frame.grid_columnconfigure(col, weight=1)

        tk.Button(action_frame, text="▶️ Run Automation", width=18,
                  command=self.run_automation,
                  bg="#92D050", fg="#222222", activebackground="#BEE395").grid(row=0, column=0, padx=15)

        tk.Button(action_frame, text="📊 Generate Summary", width=18,
                  command=self.run_summary,
                  bg="#4F81BD", fg="white", activebackground="#BEE395").grid(row=0, column=1, padx=15)

        tk.Button(action_frame, text="📈 Histogram Plot", width=18,
                  command=self.run_histogram,
                  bg="#FFC000", fg="#222222", activebackground="#BEE395").grid(row=0, column=2, padx=15)

    def create_status_box(self):
        status_frame = tk.LabelFrame(self.root, text="Status", padx=10, pady=10)
        status_frame.pack(fill="both", expand=True, padx=15, pady=10)
        container = tk.Frame(status_frame)
        container.pack(fill="both", expand=True)

        self.status_box = tk.Text(container, height=10, wrap="word", bg="white", fg="black", state="disabled")
        self.status_vsb = tk.Scrollbar(container, orient="vertical", command=self.status_box.yview)
        self.status_box.configure(yscrollcommand=self.status_vsb.set)
        self.status_box.grid(row=0, column=0, sticky="nsew")
        self.status_vsb.grid(row=0, column=1, sticky="ns")
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

    def show_status(self, message):
        self.status_box.config(state="normal")
        self.status_box.insert("end", message + "\n")
        self.status_box.config(state="disabled")

    def clear_all(self):
        try:
            # Reset path entry
            self.path_var.set("")

            # Clear status panel (Text widget)
            if hasattr(self, "status_box") and self.status_box is not None:
                self.status_box.config(state="normal")
                self.status_box.delete("1.0", "end")
                self.status_box.config(state="disabled")

            # Clear parameter listbox selection
            self.param_listbox.selection_clear(0, "end")

            # Clear plot
            self.ax.clear()
            self.ax.set_title("No Data")
            self.canvas.draw_idle()

            # Clear stats text inside figure
            self.stats_text_obj.set_text("")

            # Clear logger frame if you have one
            if hasattr(self, "logger_frame") and self.logger_frame is not None:
                self.logger_frame.config(state="normal")
                self.logger_frame.delete("1.0", "end")
                self.logger_frame.config(state="disabled")

            self.logger.log_success("Cleared all inputs, outputs, and logs.")
        except Exception as e:
            self.logger.log_error(f"Error clearing all: {e}")

    def create_exit_button(self):
        frame = tk.Frame(self.root, bg=self.bg_color)
        frame.pack(fill="x", side="bottom", padx=15, pady=5)
        tk.Button(frame, text="EXIT", width=12, bg="#d32f2f", fg="white",
                  command=self.root.destroy).pack(side="right", pady=10)
        tk.Button(frame, text="Clear All", width=12, command=self.clear_all,
                  bg="#ffcccc", fg=self.fg_color, activebackground=self.btn_active).pack(side="right", padx=10)

    # --- Actions ---
    def run_automation(self):
        try:
            wat_file = self.path_var.get()
            if not wat_file:
                self.show_status(self.logger.log_error("No file selected!"))
                return

            excel_filename = wat_file.replace(".wat", ".xlsx")

            # Parse WAT file → returns metadata
            type_no, process, pcm_spec, lot_id, date, clean_qty = self.parser.parse_file(wat_file, excel_filename)

            # Add per-unit data
            self.show_status(self.builder.add_per_unit_data(excel_filename, clean_qty))

            # Extract site values from the first worksheet
            wb = load_workbook(excel_filename)
            main_ws = wb[wb.sheetnames[0]]
            site_values = extract_site_values(main_ws, start_row=6, col="B")

            # Add per-wafer data
            self.show_status(self.builder.add_per_wafer_data(
                excel_filename,
                clean_qty,
                site_values,
                type_no,
                process,
                pcm_spec,
                lot_id,
                date
            ))

            # Final success message
            self.show_status(self.logger.log_success(f"Automation complete. Deliverables saved to {excel_filename}"))

        except Exception as e:
            self.show_status(str(e))

    def run_summary(self):
        try:
            wat_file = self.path_var.get()
            if not wat_file:
                self.show_status(self.logger.log_error("No file selected!"))
                return
            excel_filename = wat_file.replace(".wat", ".xlsx")
            self.show_status(self.builder.run_summary(excel_filename))
        except Exception as e:
            self.show_status(str(e))

    def run_histogram(self):
        try:
            wat_file = self.path_var.get()
            if not wat_file:
                self.show_status(self.logger.log_error("No file selected!"))
                return

            excel_filename = wat_file.replace(".wat", ".xlsx")

            # Launch HistogramGUI in a new window
            HistogramGUI(tk.Toplevel(self.root), excel_filename, self.logger)

            self.show_status(self.logger.log_success(f"Opened histogram viewer for {excel_filename}"))

        except Exception as e:
            self.show_status(str(e))

            
# --- Main Entry Point ---
if __name__ == "__main__":
    root = tk.Tk()
    root.title("WAT Data Automatio v1.1.1")
    root.iconbitmap(resource_path("sprout.ico"))  # subtle window icon only
    app = GuiController(root)
    root.mainloop()
    # Ensure process exits when GUI closes
    root.quit()

